from django.shortcuts import render, get_object_or_404, redirect
from .models import ProductPrice, Wilaya, Moughataa, Commune, PointDeVente, ProductType, Product, Cart, CartProduct
from .forms import ProductPriceForm, WilayaForm, MoughataaForm, CommuneForm, PointDeVenteForm, ProductTypeForm, ProductForm, CartForm, CartProductForm
import openpyxl
from django.http import HttpResponse ,HttpResponseRedirect
import os
import csv
from django.conf import settings
from django.contrib import messages
from django.urls import reverse
from datetime import datetime

def home(request):
    return render(request, 'home.html')

# Wilaya Views
def wilaya_list(request):
 wilayas = Wilaya.objects.all() 
 return render(request, 'wilaya_list.html', {'wilayas': wilayas})

def wilaya_create(request):
    if request.method == 'POST':
        form = WilayaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('wilaya_list')
    else:
        form = WilayaForm()
    return render(request, 'wilaya_form.html', {'form': form})

def wilaya_update(request, id):
    wilaya = get_object_or_404(Wilaya, id=id)
    if request.method == 'POST':
        form = WilayaForm(request.POST, instance=wilaya)
        if form.is_valid():
            form.save()
            return redirect('wilaya_list')
    else:
        form = WilayaForm(instance=wilaya)
    return render(request, 'wilaya_form.html', {'form': form})


def wilaya_delete(request, id):
    wilaya = get_object_or_404(Wilaya, id=id)
    if request.method == 'POST':
        wilaya.delete()
        return redirect('wilaya_list')
    return render(request, 'wilaya_confirm_delete.html', {'wilaya': wilaya})

#wilaya export 

def export_wilayas_to_excel(request):
    # Créer un fichier Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Wilayas"

    # Ajouter les en-têtes
    sheet.append(["ID", "Code", "Nom"])

    # Récupérer les données
    for wilaya in Wilaya.objects.all():
        sheet.append([wilaya.id, wilaya.code, wilaya.name])

    # Enregistrer dans la réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="wilayas.xlsx"'
    workbook.save(response)

    # Ajouter un message de confirmation
    messages.success(request, "Exportation réussie !")

    return response

#wilaya import 




def import_wilayas_from_excel(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]

        # Vérification du type de fichier
        if excel_file.name.endswith('.xlsx'):
            # Si c'est un fichier Excel
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # Tente de récupérer les colonnes
            headers = [cell.value for cell in sheet[1]]
            
            if "wilaya_code" in headers and "wilaya_name" in headers:
                wilaya_code_index = headers.index("wilaya_code")
                wilaya_name_index = headers.index("wilaya_name")
                
                # Lire les données à partir de la 2e ligne
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    wilaya_code = row[wilaya_code_index]
                    wilaya_name = row[wilaya_name_index]

                    # Enregistrer ou mettre à jour les wilayas
                    Wilaya.objects.update_or_create(code=wilaya_code, defaults={"name": wilaya_name})

            else:
                messages.error(request, "Le fichier Excel ne contient pas les colonnes attendues ('wilaya_code', 'wilaya_name').")
                return redirect("wilaya_list")
            
        elif excel_file.name.endswith('.csv'):
            # Si c'est un fichier CSV
            csv_file = csv.reader(excel_file.read().decode('utf-8').splitlines())
            headers = next(csv_file)  # Lire la première ligne comme les en-têtes

            if "wilaya_code" in headers and "wilaya_name" in headers:
                wilaya_code_index = headers.index("wilaya_code")
                wilaya_name_index = headers.index("wilaya_name")

                for row in csv_file:
                    wilaya_code = row[wilaya_code_index]
                    wilaya_name = row[wilaya_name_index]

                    # Enregistrer ou mettre à jour les wilayas
                    Wilaya.objects.update_or_create(code=wilaya_code, defaults={"name": wilaya_name})

            else:
                messages.error(request, "Le fichier CSV ne contient pas les colonnes attendues ('wilaya_code', 'wilaya_name').")
                return redirect("wilaya_list")
        else:
            messages.error(request, "Le fichier doit être au format .xlsx ou .csv")
            return redirect("wilaya_list")

        # Message de succès
        messages.success(request, "Importation réussie !")
        return redirect("wilaya_list")

    return render(request, "import_wilayas.html")


# Moughataa Views
def moughataa_list(request):
    moughataas = Moughataa.objects.all()
    for m in moughataas:
        print(f"Moughataa ID: {m.id}")  # Vérifier si chaque objet a bien un ID
    return render(request, 'moughataa_list.html', {'moughataas': moughataas})

def moughataa_create(request):
    if request.method == 'POST':
        form = MoughataaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('moughataa_list')
    else:
        form = MoughataaForm()
    return render(request, 'moughataa_form.html', {'form': form})

def moughataa_update(request, id):
    moughataa = get_object_or_404(Moughataa, id=id)
    if request.method == 'POST':
        form = MoughataaForm(request.POST, instance=moughataa)
        if form.is_valid():
            form.save()
            return redirect('moughataa_list')
    else:
        form = MoughataaForm(instance=moughataa)
    return render(request, 'moughataa_form.html', {'form': form})

def moughataa_delete(request, id):
    moughataa = get_object_or_404(Moughataa, id=id)
    if request.method == 'POST':
        moughataa.delete()
        return redirect('moughataa_list')
    return render(request, 'moughataa_confirm_delete.html', {'moughataa': moughataa})

#  Fonction pour exporter les Moughataas en Excel
def export_moughataas(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Moughataas"

    # Ajouter les en-têtes
    ws.append(["ID", "Code", "Label", "Wilaya ID"])

    # Remplir avec les données
    for moughataa in Moughataa.objects.all():
        ws.append([moughataa.id, moughataa.code, moughataa.label, moughataa.wilaya.id])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="moughataas.xlsx"'
    wb.save(response)

    return response

# ✅ Fonction pour importer les Moughataas depuis un fichier Excel
def import_moughataas(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        wb = openpyxl.load_workbook(file)
        ws = wb.active  # Ouvre la première feuille du fichier Excel

        for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (en-têtes)
            wilaya_name, moughataa_name, moughataa_code = row[0], row[1], row[2]

            # Trouver l'ID de la Wilaya à partir du nom
            wilaya = Wilaya.objects.filter(name=wilaya_name).first()

            if wilaya:
                # Créer ou mettre à jour la Moughataa
                Moughataa.objects.update_or_create(
                    code=moughataa_code,  # Code unique pour chaque Moughataa
                    defaults={"label": moughataa_name, "wilaya": wilaya}
                )
            else:
                # Si la Wilaya n'est pas trouvée, ignorer ou gérer cette erreur
                print(f"Wilaya '{wilaya_name}' non trouvée. La Moughataa '{moughataa_name}' ne sera pas ajoutée.")

        messages.success(request, "Importation des Moughataas réussie !")
        return redirect("moughataa_list")

    return render(request, "moughataa_import.html")


# Commune Views
def commune_list(request):
    communes = Commune.objects.all()
    return render(request, 'commune_list.html', {'communes': communes})

def commune_create(request):
    if request.method == 'POST':
        form = CommuneForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('commune_list')
    else:
        form = CommuneForm()
    return render(request, 'commune_form.html', {'form': form})

def commune_update(request, id):
    commune = get_object_or_404(Commune, id=id)
    if request.method == 'POST':
        form = CommuneForm(request.POST, instance=commune)
        if form.is_valid():
            form.save()
            return redirect('commune_list')
    else:
        form = CommuneForm(instance=commune)
    return render(request, 'commune_form.html', {'form': form})

def commune_delete(request, id):
    commune = get_object_or_404(Commune, id=id)
    if request.method == 'POST':
        commune.delete()
        return redirect('commune_list')
    return render(request, 'commune_confirm_delete.html', {'commune': commune})


# Exporter les Communes en Excel 
def export_communes(request):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="communes.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Communes"

    # Ajouter les en-têtes
    headers = ["ID", "Code", "Nom", "Moughataa"]
    sheet.append(headers)

    # Remplir avec les données
    for commune in Commune.objects.all():
        sheet.append([commune.id, commune.code, commune.name, commune.moughataa.label])

    workbook.save(response)
    return response

def import_communes(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Le fichier doit être au format .xlsx")
            return HttpResponseRedirect(reverse("commune_list"))

        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active  # ou utilisez sheet = workbook['Nom_de_votre_onglet'] si nécessaire
        
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (en-têtes)
            commune_code = row[0]  # Colonne A: Code de la commune
            commune_name = row[1]  # Colonne B: Nom de la commune
            moughataa_name = row[2]  # Colonne C: Nom de la Moughataa

            try:
                # Récupérer la Moughataa associée par son nom
                moughataa = Moughataa.objects.get(label=moughataa_name)
            except Moughataa.DoesNotExist:
                messages.error(request, f"La Moughataa '{moughataa_name}' n'existe pas.")
                continue

            # Créer ou mettre à jour la commune
            Commune.objects.update_or_create(
                code=commune_code,
                defaults={"name": commune_name, "moughataa": moughataa}
            )

        messages.success(request, "Importation des Communes réussie !")
        return HttpResponseRedirect(reverse("commune_list"))

    return render(request, "commune_import.html")


# PointDeVente Views
def pointdevente_list(request):
    pointsdevente = PointDeVente.objects.all()
    return render(request, 'pointdevente_list.html', {'pointsdevente': pointsdevente})

def pointdevente_create(request):
    if request.method == 'POST':
        form = PointDeVenteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('pointdevente_list')
    else:
        form = PointDeVenteForm()
    return render(request, 'pointdevente_form.html', {'form': form})

def pointdevente_update(request, id):
    pointdevente = get_object_or_404(PointDeVente, id=id)
    if request.method == 'POST':
        form = PointDeVenteForm(request.POST, instance=pointdevente)
        if form.is_valid():
            form.save()
            return redirect('pointdevente_list')
    else:
        form = PointDeVenteForm(instance=pointdevente)
    return render(request, 'pointdevente_form.html', {'form': form})

def pointdevente_delete(request, id):
    pointdevente = get_object_or_404(PointDeVente, id=id)
    if request.method == 'POST':
        pointdevente.delete()
        return redirect('pointdevente_list')
    return render(request, 'pointdevente_confirm_delete.html', {'pointdevente': pointdevente})


def export_pointdevente(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pointdevente.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Points de Vente"

    headers = ['Code', 'Type', 'GPS Latitude', 'GPS Longitude', 'Commune']
    sheet.append(headers)

    for point in PointDeVente.objects.all():
        sheet.append([
            point.code,
            point.type,
            point.gps_lat,
            point.gps_lon,
            point.commune.name if point.commune else ''
        ])

    workbook.save(response)
    return response

def import_pointdevente(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('pointdevente_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, type_, gps_lat, gps_lon, commune_name = row
            commune, created = Commune.objects.get_or_create(name=commune_name)

            PointDeVente.objects.create(
                code=code,
                type=type_,
                gps_lat=gps_lat,
                gps_lon=gps_lon,
                commune=commune
            )

        return redirect('pointdevente_list')

    return render(request, 'pointdevente_list.html')


# ProductType Views
def producttype_list(request):
    producttypes = ProductType.objects.all()
    return render(request, 'producttype_list.html', {'producttypes': producttypes})

def producttype_create(request):
    if request.method == 'POST':
        form = ProductTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('producttype_list')
    else:
        form = ProductTypeForm()
    return render(request, 'producttype_form.html', {'form': form})

def producttype_update(request, id):
    producttype = get_object_or_404(ProductType, id=id)
    if request.method == 'POST':
        form = ProductTypeForm(request.POST, instance=producttype)
        if form.is_valid():
            form.save()
            return redirect('producttype_list')
    else:
        form = ProductTypeForm(instance=producttype)
    return render(request, 'producttype_form.html', {'form': form})

def producttype_delete(request, id):
    producttype = get_object_or_404(ProductType, id=id)
    if request.method == 'POST':
        producttype.delete()
        return redirect('producttype_list')
    return render(request, 'producttype_confirm_delete.html', {'producttype': producttype})

def export_producttype(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=producttypes.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Types de Produits"

    headers = ['Code', 'Label', 'Description']
    sheet.append(headers)

    for producttype in ProductType.objects.all():
        sheet.append([
            producttype.code,
            producttype.label,
            producttype.description
        ])

    workbook.save(response)
    return response

def import_producttype(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('producttype_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, label, description = row
            ProductType.objects.create(
                code=code,
                label=label,
                description=description
            )

        return redirect('producttype_list')

    return render(request, 'producttype_list.html')

# Product Views
def product_list(request):
    products = Product.objects.all()
    return render(request, 'product_list.html', {'products': products})

def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'product_form.html', {'form': form})

def product_update(request, id):
    product = get_object_or_404(Product, id=id)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'product_form.html', {'form': form})

def product_delete(request, id):
    product = get_object_or_404(Product, id=id)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'product_confirm_delete.html', {'product': product})

def export_product(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Produits"

    headers = ['Code', 'Nom', 'Description', 'Unité de Mesure', 'Type de Produit']
    sheet.append(headers)

    for product in Product.objects.all():
        sheet.append([
            product.code,
            product.name,
            product.description,
            product.unit_measure,
            product.product_type.label
        ])

    workbook.save(response)
    return response

def import_product(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('product_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, name, description, unit_measure, product_type_label = row
            product_type = ProductType.objects.filter(label=product_type_label).first()

            if product_type:
                Product.objects.create(
                    code=code,
                    name=name,
                    description=description,
                    unit_measure=unit_measure,
                    product_type=product_type
                )

        return redirect('product_list')

    return render(request, 'product_list.html')

# Cart Views
def cart_list(request):
    carts = Cart.objects.all()
    return render(request, 'cart_list.html', {'carts': carts})

def cart_create(request):
    if request.method == 'POST':
        form = CartForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cart_list')
    else:
        form = CartForm()
    return render(request, 'cart_form.html', {'form': form})

def cart_update(request, id):
    cart = get_object_or_404(Cart, id=id)
    if request.method == 'POST':
        form = CartForm(request.POST, instance=cart)
        if form.is_valid():
            form.save()
            return redirect('cart_list')
    else:
        form = CartForm(instance=cart)
    return render(request, 'cart_form.html', {'form': form})

def cart_delete(request, id):
    cart = get_object_or_404(Cart, id=id)
    if request.method == 'POST':
        cart.delete()
        return redirect('cart_list')
    return render(request, 'cart_confirm_delete.html', {'cart': cart})

def export_cart(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=carts.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Paniers"

    headers = ['Code', 'Nom', 'Description']
    sheet.append(headers)

    for cart in Cart.objects.all():
        sheet.append([
            cart.code,
            cart.name,
            cart.description
        ])

    workbook.save(response)
    return response

def import_cart(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('cart_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, name, description = row

            Cart.objects.create(
                code=code,
                name=name,
                description=description
            )

        return redirect('cart_list')

    return render(request, 'cart_list.html')

# CartProduct Views
def cartproduct_list(request):
    cartproducts = CartProduct.objects.all()
    return render(request, 'cartproduct_list.html', {'cartproducts': cartproducts})

def cartproduct_create(request):
    if request.method == 'POST':
        form = CartProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cartproduct_list')
    else:
        form = CartProductForm()
    return render(request, 'cartproduct_form.html', {'form': form})

def cartproduct_update(request, id):
    cartproduct = get_object_or_404(CartProduct, id=id)
    if request.method == 'POST':
        form = CartProductForm(request.POST, instance=cartproduct)
        if form.is_valid():
            form.save()
            return redirect('cartproduct_list')
    else:
        form = CartProductForm(instance=cartproduct)
    return render(request, 'cartproduct_form.html', {'form': form})

def cartproduct_delete(request, id):
    cartproduct = get_object_or_404(CartProduct, id=id)
    if request.method == 'POST':
        cartproduct.delete()
        return redirect('cartproduct_list')
    return render(request, 'cartproduct_confirm_delete.html', {'cartproduct': cartproduct})

def export_cartproduct(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cartproducts.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Produits des Paniers"

    headers = ['Panier', 'Produit', 'Pondération', 'Date Début', 'Date Fin']
    sheet.append(headers)

    for cartproduct in CartProduct.objects.all():
        sheet.append([
            cartproduct.cart.name,
            cartproduct.product.name,
            cartproduct.weight,
            cartproduct.date_from,
            cartproduct.date_to if cartproduct.date_to else ''
        ])

    workbook.save(response)
    return response

def import_cartproduct(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('cartproduct_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cart_name, product_name, weight, date_from, date_to = row

            cart = Cart.objects.filter(name=cart_name).first()
            product = Product.objects.filter(name=product_name).first()

            if cart and product:
                CartProduct.objects.create(
                    cart=cart,
                    product=product,
                    weight=weight,
                    date_from=date_from,
                    date_to=date_to
                )

        return redirect('cartproduct_list')

    return render(request, 'cartproduct_list.html')


#prix views
# Liste des prix
def prix_list(request):
    les_prix = ProductPrice.objects.all()
    return render(request, 'prix_list.html', {'les_prix': les_prix})


# Ajouter un prix
def prix_create(request):
    if request.method == 'POST':
        form = ProductPriceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('prix_list')
    else:
        form = ProductPriceForm()
    return render(request, 'prix_form.html', {'form': form})

# Modifier un prix
def prix_update(request, id):
    prix = get_object_or_404(ProductPrice, id=id)
    if request.method == 'POST':
        form = ProductPriceForm(request.POST, instance=prix)
        if form.is_valid():
            form.save()
            return redirect('prix_list')
    else:
        form = ProductPriceForm(instance=prix)
    return render(request, 'prix_form.html', {'form': form})

# Supprimer un prix
def prix_delete(request, id):
    prix = get_object_or_404(ProductPrice, id=id)
    if request.method == 'POST':
        prix.delete()
        return redirect('prix_list')
    return render(request, 'prix_confirm_delete.html', {'prix': prix})

def structures_list(request):
    wilayas = Wilaya.objects.all()
    moughataas = Moughataa.objects.all()
    communes = Commune.objects.all()
    return render(request, 'structures_list.html', {
        'wilayas': wilayas,
        'moughataas': moughataas,
        'communes': communes,
    })

def export_productprice(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productprices.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Prix des Produits"

    headers = ['Produit', 'Point de Vente', 'Valeur', 'Date Début', 'Date Fin']
    sheet.append(headers)

    for productprice in ProductPrice.objects.all():
        sheet.append([
            productprice.product.name,
            productprice.point_of_sale.code,
            productprice.value,
            productprice.date_from,
            productprice.date_to if productprice.date_to else ''
        ])

    workbook.save(response)
    return response

def import_productprice(request):
    if request.method == "POST":
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            return redirect('productprice_list')

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            product_name, point_of_sale_code, value, date_from, date_to = row

            product = Product.objects.filter(name=product_name).first()
            point_of_sale = PointDeVente.objects.filter(code=point_of_sale_code).first()

            if product and point_of_sale:
                ProductPrice.objects.create(
                    product=product,
                    point_of_sale=point_of_sale,
                    value=value,
                    date_from=date_from,
                    date_to=date_to
                )

        return redirect('productprice_list')

    return render(request, 'productprice_list.html')

##calcul de L'INPC

def calculate_inpc(request):
    if request.method == 'POST':
        # Récupérer la date (mois et année)
        year = request.POST.get('year')
        month = request.POST.get('month')
        date_from = f"{year}-{month}-01"
        date_to = f"{year}-{month}-31"  # Assuming the month has 31 days, adjust for the month
        
        # Récupérer les paniers et les produits associés
        carts = Cart.objects.all()
        total_weighted_price = 0
        total_weight = 0

        for cart in carts:
            cart_products = CartProduct.objects.filter(cart=cart)
            
            for cart_product in cart_products:
                product = cart_product.product
                weight = cart_product.weight
                
                # Récupérer les prix du produit pour le mois sélectionné
                product_prices = ProductPrice.objects.filter(
                    product=product,
                    date_from__lte=date_to,
                    date_to__gte=date_from
                )
                
                # Calculer le prix moyen du produit pour ce mois
                total_price = 0
                count = 0
                for price in product_prices:
                    total_price += price.value
                    count += 1
                
                if count > 0:
                    avg_price = total_price / count
                else:
                    avg_price = 0

                # Calculer le prix moyen pondéré du produit pour le panier
                total_weighted_price += weight * avg_price
                total_weight += weight

        if total_weight > 0:
            inpc = total_weighted_price / total_weight
        else:
            inpc = 0
        
        return render(request, 'inpc_result.html', {'inpc': inpc})

    return render(request, 'calculate_inpc.html')